import streamlit as st
import pandas as pd
import yfinance as yf
import requests
import io
import re
import time
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go  
from plotly.subplots import make_subplots 
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==========================================
# 0. 初始化設定與共用函數 (加入連動 Callback)
# ==========================================
if 'analyzed_input' not in st.session_state:
    st.session_state.analyzed_input = None
if 'target_date' not in st.session_state:
    st.session_state.target_date = None
if 'period_choice' not in st.session_state:
    st.session_state.period_choice = "近三個月 (64天)"

# 🟢 建立兩個開關的連動同步函數
def sync_period_top():
    st.session_state.period_choice = st.session_state.radio_top

def sync_period_bottom():
    st.session_state.period_choice = st.session_state.radio_bottom

st.set_page_config(page_title="牧場小霸王", page_icon="📈", layout="wide")

@st.cache_data(ttl=3600, show_spinner=False)
def get_latest_trading_date():
    try:
        idx_data = yf.Ticker("^TWII").history(period="5d")
        return idx_data.index[-1].strftime("%Y/%m/%d")
    except:
        return datetime.now().strftime("%Y/%m/%d")
        
@st.cache_data
def load_stock_list():
    try:
        df = pd.read_excel('TW50100.xlsx', engine='openpyxl', dtype=str)
        return {str(row[df.columns[1]]): str(row[df.columns[0]]).replace('.0', '') for _, row in df.iterrows()}, True
    except: return {}, False

# ==========================================
# 副程式 1：抓取 YFinance 資料
# ==========================================
def step1_fetch_yf_data(ticker, raw_ticker, auto_fallback, target_date_str):
    end_dt = pd.to_datetime(target_date_str, format='%Y/%m/%d') + pd.Timedelta(days=1)
    start_dt = end_dt - pd.DateOffset(months=10) 
    
    start_str = start_dt.strftime('%Y-%m-%d')
    end_str = end_dt.strftime('%Y-%m-%d')

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    })

    hist = pd.DataFrame()
    
    try:
        temp_hist = yf.Ticker(ticker, session=session).history(start=start_str, end=end_str)
        if not temp_hist.empty:
            hist = temp_hist[temp_hist['Volume'] > 0]
    except Exception:
        pass 

    if hist.empty and auto_fallback and raw_ticker:
        ticker_two = f"{raw_ticker}.TWO"
        try:
            temp_hist_two = yf.Ticker(ticker_two, session=session).history(start=start_str, end=end_str)
            if not temp_hist_two.empty:
                hist_two = temp_hist_two[temp_hist_two['Volume'] > 0]
                if not hist_two.empty:
                    hist = hist_two
                    ticker = ticker_two
        except Exception:
            pass

    return hist, ticker

# ==========================================
# 副程式 2：產生 K線、均線、KD、MACD 與 一目均衡表
# ==========================================
def step2_calc_tech_indicators(hist):
    df = hist.copy()
    
    df['MA5'] = df['Close'].rolling(window=5).mean()
    df['MA10'] = df['Close'].rolling(window=10).mean()
    df['MA20'] = df['Close'].rolling(window=20).mean()

    low_min = df['Low'].rolling(window=9).min()
    high_max = df['High'].rolling(window=9).max()
    rsv = (df['Close'] - low_min) / (high_max - low_min + 1e-9) * 100
    df['K'] = rsv.ewm(com=2, adjust=False).mean()
    df['D'] = df['K'].ewm(com=2, adjust=False).mean()
    
    ema12 = df['Close'].ewm(span=12, adjust=False).mean()
    ema26 = df['Close'].ewm(span=26, adjust=False).mean()
    df['DIF'] = ema12 - ema26
    df['MACD'] = df['DIF'].ewm(span=9, adjust=False).mean()
    df['OSC'] = (df['DIF'] - df['MACD']) * 2 
    df['Volume'] = df['Volume'] / 1000  
    
    high_9 = df['High'].rolling(window=9).max()
    low_9 = df['Low'].rolling(window=9).min()
    df['Tenkan'] = (high_9 + low_9) / 2  

    high_26 = df['High'].rolling(window=26).max()
    low_26 = df['Low'].rolling(window=26).min()
    df['Kijun'] = (high_26 + low_26) / 2 

    high_52 = df['High'].rolling(window=52).max()
    low_52 = df['Low'].rolling(window=52).min()
    
    senkou_a_unaligned = (df['Tenkan'] + df['Kijun']) / 2 
    senkou_b_unaligned = (high_52 + low_52) / 2           
    df['Chikou'] = df['Close'].shift(-26)                 

    if not df.empty:
        last_date = df.index[-1]
        future_dates = pd.bdate_range(start=last_date + pd.Timedelta(days=1), periods=26)
        future_df = pd.DataFrame(index=future_dates)
        
        df_extended = pd.concat([df, future_df])
        
        df_extended['Senkou_A'] = senkou_a_unaligned.reindex(df_extended.index).shift(26)
        df_extended['Senkou_B'] = senkou_b_unaligned.reindex(df_extended.index).shift(26)
    else:
        df_extended = df

    return df_extended

# ==========================================
# 副程式 3：產生分價量統計與圖表
# ==========================================
def step3_process_volume_profile(valid_hist):
    current_price_round = round(valid_hist['Close'].dropna().iloc[-1], 2)
    max_price, min_price = valid_hist['High'].max(), valid_hist['Low'].min()
    if max_price == min_price:
        max_price, min_price = min_price * 1.05, min_price * 0.95
    
    bin_size = (max_price - min_price) / 20
    curr_bin_idx = 19 if current_price_round >= max_price else (0 if current_price_round <= min_price else min(19, int((current_price_round - min_price) / bin_size)))
    
    bins_data = [{'idx': i, 'start': min_price + i * bin_size, 'end': min_price + (i + 1) * bin_size, 'mid': (min_price + i * bin_size + min_price + (i + 1) * bin_size) / 2, 'label': f"{min_price + i * bin_size:.2f} ~ {min_price + (i + 1) * bin_size:.2f}", 'disp_label': f"{'** ' if i == curr_bin_idx else ''}{min_price + i * bin_size:.2f} ~ {min_price + (i + 1) * bin_size:.2f}", 'is_current': (i == curr_bin_idx), 'vol': 0} for i in range(20)]
    
    all_price_vols = []
    for _, row in valid_hist.iterrows():
        o, h, l, c, v = round(row['Open'], 2), round(row['High'], 2), round(row['Low'], 2), round(row['Close'], 2), row['Volume']
        if l > h: l, h = h, l 
        vol_open, vol_close, vol_dist_total = v * 0.05, v * 0.30, v * 0.65
        
        ticks, curr = [], l
        while curr <= h:
            ticks.append(curr)
            ts = 0.01 if curr < 10 else (0.05 if curr < 50 else (0.1 if curr < 100 else (0.5 if curr < 500 else (1.0 if curr < 1000 else 5.0))))
            curr = round(curr + ts, 2)
        
        n_ticks = len(ticks)
        vol_per_tick = vol_dist_total / n_ticks if n_ticks > 0 else 0
        all_price_vols.extend([{'Price': o, 'Vol': vol_open}, {'Price': c, 'Vol': vol_close}] + [{'Price': t, 'Vol': vol_per_tick} for t in ticks])
            
    df_all_vols = pd.DataFrame(all_price_vols)
    for price, vol in df_all_vols.groupby('Price')['Vol'].sum().items():
        if price >= max_price: bins_data[-1]['vol'] += vol
        elif price <= min_price: bins_data[0]['vol'] += vol
        else: bins_data[min(19, int((price - min_price) / bin_size))]['vol'] += vol
            
    all_intervals_disp = sorted(bins_data, key=lambda x: x['idx'], reverse=True)
    
    df_plot = pd.DataFrame({
        '價格區間': [item['label'] for item in all_intervals_disp],
        '累積成交量 (張)': [int(item['vol']) for item in all_intervals_disp],
        '標記': ['現價所在' if item['is_current'] else '一般區間' for item in all_intervals_disp]
    })
    fig_vol = px.bar(df_plot, x='累積成交量 (張)', y='價格區間', color='標記', color_discrete_map={'現價所在': '#FF4B4B', '一般區間': '#60B4FF'}, orientation='h')
    fig_vol.update_yaxes(categoryorder='array', categoryarray=df_plot['價格區間'])
    fig_vol.update_layout(yaxis=dict(title="價格區間", autorange="reversed"), margin=dict(l=0, r=0, t=30, b=0), height=500)
    
    return bins_data, all_intervals_disp, fig_vol, current_price_round

# ==========================================
# 副程式 4：關鍵分價量支撐
# ==========================================
def step4_find_support_resistance(bins_data, current_price_round):
    top_5_above = sorted(sorted([b for b in bins_data if b['mid'] >= current_price_round and b['vol'] > 0], key=lambda x: x['vol'], reverse=True)[:5], key=lambda x: x['start'], reverse=True)
    top_5_below = sorted(sorted([b for b in bins_data if b['mid'] < current_price_round and b['vol'] > 0], key=lambda x: x['vol'], reverse=True)[:5], key=lambda x: x['start'], reverse=True)
    return top_5_above, top_5_below

# ==========================================
# 副程式 5 & 6 相關：即時下載外資、投信 CSV 與 融資券 JSON
# ==========================================
@st.cache_data(ttl=86400, show_spinner=False)
def download_twse_csv_text(date_str, inst_type):
    url = f"https://www.twse.com.tw/rwd/zh/fund/{inst_type}?date={date_str}&response=csv"
    time.sleep(1)
    try:
        res = requests.get(url, timeout=5)
        res.encoding = 'big5'
        if len(res.text) > 100:
            return res.text
        return ""
    except:
        return ""

def fetch_twse_csv_data(date_str, inst_type):
    csv_text = download_twse_csv_text(date_str, inst_type)
    if csv_text:
        try:
            return pd.read_csv(io.StringIO(csv_text), names=list(range(20)), on_bad_lines='skip')
        except: pass
    return pd.DataFrame()

@st.cache_data(ttl=86400, show_spinner=False)
def download_twse_margin_json(date_str):
    url = f"https://www.twse.com.tw/exchangeReport/MI_MARGN?response=json&date={date_str}&selectType=ALL"
    time.sleep(1) 
    try:
        res = requests.get(url, timeout=5).json()
        if res.get('stat') == 'OK':
            return res
    except: pass
    return {}

def fetch_margin_json_data(date_str, raw_ticker):
    res = download_twse_margin_json(date_str)
    if not res: return 0, 0, 0, 0
    
    tables = res.get('tables', [])
    if not tables and 'data' in res:
        tables = [{'data': res['data']}]
        
    for table in tables:
        for row in table.get('data', []):
            if str(row[0]).strip() == raw_ticker:
                m_prev = int(str(row[5]).replace(',', ''))
                m_today = int(str(row[6]).replace(',', ''))
                s_prev = int(str(row[11]).replace(',', ''))
                s_today = int(str(row[12]).replace(',', ''))
                return (m_today - m_prev), m_today, (s_today - s_prev), s_today
    return 0, 0, 0, 0

@st.cache_data(ttl=86400, show_spinner=False)
def download_tpex_csv_text(date_str, inst_type, search_type="buy"):
    url = f"https://www.tpex.org.tw/www/zh-tw/insti/{inst_type}?type=Daily&date={date_str}&searchType={search_type}&id=&response=csv"
    time.sleep(1) 
    try:
        res = requests.get(url, timeout=5, verify=False)
        if len(res.content) > 100: 
            return res.content.decode('cp950', errors='ignore')
    except: pass
    return ""

@st.cache_data(ttl=86400, show_spinner=False)
def download_tpex_margin_json(roc_date_str):
    url = f"https://www.tpex.org.tw/web/stock/margin_trading/margin_balance/margin_bal_result.php?l=zh-tw&o=json&d={roc_date_str}"
    time.sleep(1) 
    try:
        return requests.get(url, timeout=5, verify=False).json()
    except: pass
    return {}

def fetch_tpex_margin_json_data(roc_date_str, raw_ticker):
    res = download_tpex_margin_json(roc_date_str)
    if not res: return 0, 0, 0, 0
    
    tables = res.get('tables', [])
    for table in tables:
        for row in table.get('data', []):
            if str(row[0]).strip() == raw_ticker:
                m_prev = int(str(row[2]).replace(',', ''))
                m_today = int(str(row[6]).replace(',', ''))
                s_prev = int(str(row[10]).replace(',', ''))
                s_today = int(str(row[14]).replace(',', ''))
                return (m_today - m_prev), m_today, (s_today - s_prev), s_today
    return 0, 0, 0, 0

# ==========================================
# 籌碼主迴圈 
# ==========================================
def step6_extract_institutional_data(raw_ticker, valid_hist, is_otc):
    last_20_dates = valid_hist.index[-20:]
    last_10_dates = valid_hist.index[-10:]
    
    foreign_records = []
    trust_records = []
    margin_records = []
    
    def safe_parse_int(val_str):
        s = str(val_str).strip()
        if not s: return 0
        
        is_negative = False
        if s.startswith('-') or s.startswith('－') or s.startswith('−') or \
           (s.startswith('(') and s.endswith(')')) or '△' in s or '▲' in s:
            is_negative = True
            
        cleaned = re.sub(r'\D', '', s)
        if not cleaned: return 0
        
        val = int(cleaned)
        return -val if is_negative else val
    
    for d in last_20_dates:
        date_disp_str = d.strftime('%m/%d')
        
        if not is_otc:
            date_api_str = d.strftime('%Y%m%d')
            
            df_foreign = fetch_twse_csv_data(date_api_str, "TWT38U")
            net_f = 0
            if not df_foreign.empty:
                df_foreign[1] = df_foreign[1].astype(str).str.replace(r'[=" ]', '', regex=True)
                target_row = df_foreign[df_foreign[1] == raw_ticker]
                if not target_row.empty:
                    net_f = round(safe_parse_int(target_row.iloc[0, 5]) / 1000)
            foreign_records.append({'日期': date_disp_str, '外資買賣超(張)': net_f})
            
            df_trust = fetch_twse_csv_data(date_api_str, "TWT44U")
            net_t = 0
            if not df_trust.empty:
                df_trust[1] = df_trust[1].astype(str).str.replace(r'[=" ]', '', regex=True)
                target_row = df_trust[df_trust[1] == raw_ticker]
                if not target_row.empty:
                    net_t = round(safe_parse_int(target_row.iloc[0, 5]) / 1000)
            trust_records.append({'日期': date_disp_str, '投信買賣超(張)': net_t})
            
            if d in last_10_dates:
                m_change, m_today, s_change, s_today = fetch_margin_json_data(date_api_str, raw_ticker)
                margin_records.append({'日期': date_disp_str, '融資變動(張)': m_change, '融資餘額(張)': m_today, '融券變動(張)': s_change, '融券餘額(張)': s_today})
                
        else:
            date_tpex_csv_str = d.strftime('%Y/%m/%d')
            
            net_f = 0
            found_f = False
            
            csv_f_buy = download_tpex_csv_text(date_tpex_csv_str, "qfiiStat", "buy")
            if csv_f_buy:
                df_f_buy = pd.read_csv(io.StringIO(csv_f_buy), names=list(range(20)), on_bad_lines='skip')
                df_f_buy[1] = df_f_buy[1].astype(str).str.replace(r'[=" ]', '', regex=True)
                target_row = df_f_buy[df_f_buy[1] == raw_ticker]
                if not target_row.empty:
                    net_f = safe_parse_int(target_row.iloc[0, 5])
                    found_f = True
            
            if not found_f:
                csv_f_sell = download_tpex_csv_text(date_tpex_csv_str, "qfiiStat", "sell")
                if csv_f_sell:
                    df_f_sell = pd.read_csv(io.StringIO(csv_f_sell), names=list(range(20)), on_bad_lines='skip')
                    df_f_sell[1] = df_f_sell[1].astype(str).str.replace(r'[=" ]', '', regex=True)
                    target_row = df_f_sell[df_f_sell[1] == raw_ticker]
                    if not target_row.empty:
                        val = safe_parse_int(target_row.iloc[0, 5])
                        net_f = -abs(val)
                        
            foreign_records.append({'日期': date_disp_str, '外資買賣超(張)': net_f})
            
            net_t = 0
            found_t = False
            
            csv_t_buy = download_tpex_csv_text(date_tpex_csv_str, "sitcStat", "buy")
            if csv_t_buy:
                df_t_buy = pd.read_csv(io.StringIO(csv_t_buy), names=list(range(20)), on_bad_lines='skip')
                df_t_buy[1] = df_t_buy[1].astype(str).str.replace(r'[=" ]', '', regex=True)
                target_row = df_t_buy[df_t_buy[1] == raw_ticker]
                if not target_row.empty:
                    net_t = safe_parse_int(target_row.iloc[0, 5])
                    found_t = True
                    
            if not found_t:
                csv_t_sell = download_tpex_csv_text(date_tpex_csv_str, "sitcStat", "sell")
                if csv_t_sell:
                    df_t_sell = pd.read_csv(io.StringIO(csv_t_sell), names=list(range(20)), on_bad_lines='skip')
                    df_t_sell[1] = df_t_sell[1].astype(str).str.replace(r'[=" ]', '', regex=True)
                    target_row = df_t_sell[df_t_sell[1] == raw_ticker]
                    if not target_row.empty:
                        val = safe_parse_int(target_row.iloc[0, 5])
                        net_t = -abs(val)
                        
            trust_records.append({'日期': date_disp_str, '投信買賣超(張)': net_t})
            
            if d in last_10_dates:
                roc_date_str = f"{d.year - 1911}/{d.strftime('%m/%d')}"
                m_change, m_today, s_change, s_today = fetch_tpex_margin_json_data(roc_date_str, raw_ticker)
                margin_records.append({'日期': date_disp_str, '融資變動(張)': m_change, '融資餘額(張)': m_today, '融券變動(張)': s_change, '融券餘額(張)': s_today})

    df_f_res = pd.DataFrame(foreign_records)
    df_t_res = pd.DataFrame(trust_records)
    df_m_res = pd.DataFrame(margin_records)
    
    fig_f = px.bar(df_f_res, x='日期', y='外資買賣超(張)', title='近20日外資買賣超狀況', text_auto=True)
    fig_f.update_traces(marker_color=['#FF4B4B' if val > 0 else '#00B050' for val in df_f_res['外資買賣超(張)']])
    fig_f.update_layout(margin=dict(l=20, r=20, t=40, b=20), height=300)
    
    fig_t = px.bar(df_t_res, x='日期', y='投信買賣超(張)', title='近20日投信買賣超狀況', text_auto=True)
    fig_t.update_traces(marker_color=['#FF4B4B' if val > 0 else '#00B050' for val in df_t_res['投信買賣超(張)']])
    fig_t.update_layout(margin=dict(l=20, r=20, t=40, b=20), height=300)

    return df_f_res, df_t_res, df_m_res, fig_f, fig_t

# ==========================================
# 介面繪製輔助函數 (Tech Chart)
# ==========================================
def render_tech_chart(hist_extended, show_ma5, show_ma10, show_ma20, show_ichimoku, allow_zoom, lookback):
    date_strings = hist_extended.index.strftime('%Y-%m-%d')
    
    fig_k = make_subplots(rows=4, cols=1, shared_xaxes=True, vertical_spacing=0.03, 
                          row_heights=[0.4, 0.2, 0.2, 0.2], 
                          subplot_titles=("價格與均線", "KD (9,3,3)", "MACD (12,26,9)", f"成交量與{lookback}日均量"))
    
    fig_k.add_trace(go.Candlestick(x=date_strings, open=hist_extended['Open'], high=hist_extended['High'], low=hist_extended['Low'], close=hist_extended['Close'], name='K線', increasing_line_color='#FF4B4B', increasing_fillcolor='#FF4B4B', decreasing_line_color='#00B050', decreasing_fillcolor='#00B050'), row=1, col=1)
    if show_ma5: fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['MA5'], name='5MA', line=dict(color='#7A431D', width=1.5)), row=1, col=1)
    if show_ma10: fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['MA10'], name='10MA', line=dict(color='#00E5FF', width=1.5)), row=1, col=1)
    if show_ma20: fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['MA20'], name='20MA', line=dict(color='#0D47A1', width=1.5)), row=1, col=1)
    
    if show_ichimoku:
        fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['Tenkan'], name='轉換線(9)', line=dict(color='#E91E63', width=1.5)), row=1, col=1)
        fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['Kijun'], name='基準線(26)', line=dict(color='#9C27B0', width=1.5)), row=1, col=1)
        fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['Chikou'], name='遲行跨度', line=dict(color='#8D6E63', width=1.2, dash='dot')), row=1, col=1)
        fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['Senkou_A'], name='先行跨度A', line=dict(color='rgba(255, 152, 0, 0.6)', width=1)), row=1, col=1)
        fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['Senkou_B'], name='先行跨度B', fill='tonexty', fillcolor='rgba(255, 152, 0, 0.15)', line=dict(color='rgba(120, 144, 156, 0.6)', width=1)), row=1, col=1)

    fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['K'], name='K值', line=dict(color='#FF9900', width=1.2)), row=2, col=1)
    fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['D'], name='D值', line=dict(color='#0066FF', width=1.2)), row=2, col=1)
    
    macd_colors = ['#FF4B4B' if val > 0 else '#00B050' for val in hist_extended['OSC']]
    fig_k.add_trace(go.Bar(x=date_strings, y=hist_extended['OSC'], name='OSC', marker_color=macd_colors), row=3, col=1)
    fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['DIF'], name='DIF', line=dict(color='#FF9900', width=1.2)), row=3, col=1)
    fig_k.add_trace(go.Scatter(x=date_strings, y=hist_extended['MACD'], name='MACD', line=dict(color='#0066FF', width=1.2)), row=3, col=1)
    
    vol_colors = ['#FF4B4B' if row['Close'] >= row['Open'] else '#00B050' for idx, row in hist_extended.iterrows()]
    fig_k.add_trace(go.Bar(x=date_strings, y=hist_extended['Volume'], name='成交量(張)', marker_color=vol_colors), row=4, col=1)
    
    avg_vol = hist_extended['Volume'].dropna().mean()
    fig_k.add_trace(go.Scatter(x=date_strings, y=[avg_vol]*len(hist_extended), name=f'{lookback}日均量({int(avg_vol)}張)', mode='lines', line=dict(color='#FFD700', width=2, dash='dash')), row=4, col=1)
    
    fig_k.update_layout(
        xaxis=dict(type='category', visible=False), 
        xaxis2=dict(type='category', visible=False), 
        xaxis3=dict(type='category', visible=False), 
        xaxis4=dict(type='category', visible=True, title="交易日期", nticks=15),
        yaxis=dict(visible=False), 
        yaxis2=dict(visible=True), 
        yaxis3=dict(visible=True),
        yaxis4=dict(visible=True), 
        xaxis_rangeslider_visible=False, 
        margin=dict(l=4, r=4, t=30, b=4), 
        height=850, 
        hovermode='x unified', 
        showlegend=False
    )
    
    fig_k.update_xaxes(fixedrange=not allow_zoom)
    fig_k.update_yaxes(fixedrange=not allow_zoom)
    
    return fig_k

# ==========================================
# 🚀 系統主程式 (Main Program)
# ==========================================
st.title("📊 牧場小霸王")
st.markdown("支援 **技術K線均線**、**一目均衡雲帶**、**分價量防守** 與 **三大法人/融洪券籌碼分析**")

name_to_ticker, list_loaded = load_stock_list()
if not list_loaded: st.warning("⚠️ 找不到 'TW50100.xlsx'，請直接輸入股票代號。")

user_input = st.text_input("🔍 請輸入個股名稱或代號：", placeholder="例如: 台積電 或 2330")

default_date = get_latest_trading_date()
target_date_input = st.text_input("📅 請輸入查詢基準日 (西元年/月/日)：", value=default_date, placeholder="例如: 2024/01/01")

if st.button("🚀 開始分析", use_container_width=True):
    input_date_str = target_date_input.strip()
    if not input_date_str: input_date_str = default_date
        
    try:
        datetime.strptime(input_date_str, "%Y/%m/%d")
        st.session_state.analyzed_input = user_input
        st.session_state.target_date = input_date_str
    except ValueError:
        st.error("⚠️ 日期格式錯誤！請輸入正確的「西元年/月/日」格式，例如：2024/01/01")
        st.stop() 

# ------------------------------------
if st.session_state.analyzed_input:
    current_target = st.session_state.analyzed_input
    
    matched_names = [name for name in name_to_ticker.keys() if current_target in name] if list_loaded else []
    
    if len(matched_names) > 1 and current_target in matched_names:
        matched_names = [current_target] 

    if len(matched_names) == 0:
        target_name = f"自訂代號 ({current_target})"
        auto_fallback = False if current_target.lower().endswith(('.tw', '.two')) else True
        raw_ticker = current_target.split('.')[0]
        yf_ticker = current_target.upper() if not auto_fallback else f"{raw_ticker}.TW"
    elif len(matched_names) > 1:
        st.error(f"⚠️ 找到多檔股票，請輸入更明確的名稱：{', '.join(matched_names)}")
        st.stop()
    else:
        target_name = matched_names[0]
        raw_ticker = name_to_ticker[target_name]
        yf_ticker, auto_fallback = f"{raw_ticker}.TW", True

    with st.spinner('📡 正在運算核心技術指標與分價量...'):
        hist, yf_ticker = step1_fetch_yf_data(yf_ticker, raw_ticker, auto_fallback, st.session_state.target_date)
        if hist.empty:
            st.error("❌ 無法取得該日期之前的歷史資料。請確認代號與日期。")
            st.stop()
            
        hist_extended_full = step2_calc_tech_indicators(hist)
        valid_hist_full = hist_extended_full.dropna(subset=['Volume'])
        latest = valid_hist_full.iloc[-1]
        current_price_round_full = round(latest['Close'], 2)

    actual_last_date = valid_hist_full.index[-1].strftime('%Y/%m/%d')
    st.success(f"✅ {target_name} ({yf_ticker}) 分析完成！實際查詢基準日: **{actual_last_date}** / 股價: **{current_price_round_full:.2f}**")

    # ==========================================
    # 🟢 雙向連動開關 (上方位置 - 綁定 radio_top)
    # ==========================================
    st.write("")
    st.radio("🗓️ **選擇分析區間 (將同步影響下方 K 線圖與分價量)：**", 
             ["近三個月 (64天)", "近六個月 (128天)"], 
             key="radio_top",
             index=0 if st.session_state.period_choice == "近三個月 (64天)" else 1,
             on_change=sync_period_top,
             horizontal=True)
             
    lookback = 64 if "64" in st.session_state.period_choice else 128
    
    valid_hist = valid_hist_full.tail(lookback)
    hist_extended = hist_extended_full.tail(lookback + 26)
    
    bins_data, all_intervals_disp, fig_vol, current_price_round = step3_process_volume_profile(valid_hist)
    top_5_above, top_5_below = step4_find_support_resistance(bins_data, current_price_round)

    st.subheader("📊 技術指標參考")
    
    # 🟢 預先計算一目均衡表狀態
    # 1. 雲帶位置計算
    cloud_top = max(latest['Senkou_A'], latest['Senkou_B']) if not pd.isna(latest['Senkou_A']) else 0
    cloud_bottom = min(latest['Senkou_A'], latest['Senkou_B']) if not pd.isna(latest['Senkou_A']) else 0
    
    if pd.isna(latest['Senkou_A']):
        kumo_status = "⚪ 數據不足"
    elif latest['Close'] > cloud_top:
        kumo_status = "✅ 雲上 (強勢)"
    elif latest['Close'] < cloud_bottom:
        kumo_status = "⚠️ 雲下 (弱勢)"
    else:
        kumo_status = "⭕ 雲中 (盤整)"
        
    # 2. 轉基交叉
    tenkan_kijun_status = "✅ 多" if latest['Tenkan'] > latest['Kijun'] else "⚠️ 空"
    
    # 3. 遲行確認 (今天收盤價 vs 26天前的收盤價)
    past_close = valid_hist_full['Close'].iloc[-27] if len(valid_hist_full) > 26 else latest['Close']
    chikou_status = "✅ 多" if latest['Close'] > past_close else "⚠️ 空"

    # 🟢 綜合輸出表格
    st.table(pd.DataFrame({
        "項目": [
            "均線狀況", "KD狀況", "MACD狀況", 
            "一目 - 雲帶位置", "一目 - 轉基交叉", "一目 - 遲行確認"
        ],
        "狀態": [
            "✅ 多頭" if latest['MA5'] > latest['MA10'] > latest['MA20'] else ("⚠️ 空頭" if latest['MA5'] < latest['MA10'] < latest['MA20'] else "⭕ 盤整"), 
            "✅ 多" if latest['K'] > latest['D'] else "⚠️ 空", 
            "✅ 多" if latest['DIF'] > latest['MACD'] else "⚠️ 空",
            kumo_status,
            tenkan_kijun_status,
            chikou_status
        ],
        "數值細項": [
            f"5MA:{latest['MA5']:.1f} / 10MA:{latest['MA10']:.1f}", 
            f"K:{latest['K']:.1f} / D:{latest['D']:.1f}", 
            f"DIF:{latest['DIF']:.1f} / MACD:{latest['MACD']:.1f}",
            f"現價:{latest['Close']:.2f} / 雲頂:{cloud_top:.2f} / 雲底:{cloud_bottom:.2f}",
            f"轉換:{latest['Tenkan']:.2f} / 基準:{latest['Kijun']:.2f}",
            f"現價:{latest['Close']:.2f} / 26日前收盤:{past_close:.2f}"
        ]
    }))

    allow_zoom = st.checkbox("🔍 啟用圖表縮放與拖曳", value=False)
    with st.container(border=True):
        st.subheader("📈 技術分析綜合儀表板")
        c1, c2, c3, c4 = st.columns(4)
        show_ma5 = c1.checkbox("顯示 5MA", value=False)
        show_ma10 = c2.checkbox("顯示 10MA", value=True)
        show_ma20 = c3.checkbox("顯示 20MA", value=False)
        show_ichimoku = c4.checkbox("☁️ 顯示一目均衡表 (雲帶)", value=False)
        
        fig_tech = render_tech_chart(hist_extended, show_ma5, show_ma10, show_ma20, show_ichimoku, allow_zoom, lookback)
        st.plotly_chart(fig_tech, use_container_width=True)

    st.divider()

    # ==========================================
    # 🟢 雙向連動開關 (下方位置 - 綁定 radio_bottom)
    # ==========================================
    st.radio("🗓️ **快速切換分價量統計區間 (與上方開關完全同步)：**", 
             ["近三個月 (64天)", "近六個月 (128天)"], 
             key="radio_bottom",
             index=0 if st.session_state.period_choice == "近三個月 (64天)" else 1,
             on_change=sync_period_bottom,
             horizontal=True)

    st.subheader("📏 均線落點分價區間")
    col_ma1, col_ma2, col_ma3 = st.columns(3)
    
    ma_settings = [
        (col_ma1, "5MA", latest['MA5']), 
        (col_ma2, "10MA", latest['MA10']), 
        (col_ma3, "20MA", latest['MA20'])
    ]
    
    for col, ma_name, ma_val in ma_settings:
        with col:
            if pd.isna(ma_val): 
                st.write(f"**{ma_name}**：無資料")
                continue
                
            target_bin = None
            for b in all_intervals_disp:
                if b['start'] <= ma_val <= b['end']:
                    target_bin = b
                    break
                    
            if not target_bin:
                if ma_val > all_intervals_disp[0]['end']: 
                    target_bin = all_intervals_disp[0] 
                else: 
                    target_bin = all_intervals_disp[-1] 
            
            st.markdown(f"**{ma_name} ({ma_val:.2f})**")
            st.write(f"落於區間：`{target_bin['label']}`")
            st.write(f"區間籌碼：**{int(target_bin['vol']):,}** 張")
    
    st.subheader(f"📊 {lookback}日分價量參考圖")
    fig_vol.update_xaxes(fixedrange=not allow_zoom)
    fig_vol.update_yaxes(fixedrange=not allow_zoom)
    st.plotly_chart(fig_vol, use_container_width=True)

    st.subheader("🎯 關鍵支撐與壓力 (Top 5)")
    col1, col2 = st.columns(2)
    with col1:
        st.write("**⬆️ 向上方壓力區**")
        for item in top_5_above: st.write(f"`{item['disp_label']:<20}` | **{int(item['vol']):,}** 張")
    with col2:
        st.write("**⬇️ 向下方支撐區**")
        for item in top_5_below: st.write(f"`{item['disp_label']:<20}` | **{int(item['vol']):,}** 張")

    st.divider()
    show_debug = False 
    
    c_title, c_btn = st.columns([4, 1])
    with c_title:
        st.subheader("📈 近期市場籌碼動向 (外資投信 20日 / 融資券 10日)")
    with c_btn:
        st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
        if st.button("🔄 重新下載快取", use_container_width=True):
            download_twse_csv_text.clear()
            download_tpex_csv_text.clear()
            download_twse_margin_json.clear()
            download_tpex_margin_json.clear()
            
    is_otc = yf_ticker.endswith('.TWO')
    
    if show_debug and is_otc:
        st.warning("🔍 【Debug 模式】目前為上櫃股票，以下印出最新一天的原始純文字 (已關閉表格渲染以防止當機)：")
        test_d = valid_hist.index[-1]
        test_tpex_csv = test_d.strftime('%Y/%m/%d') 
        test_roc = f"{test_d.year - 1911}/{test_d.strftime('%m/%d')}"
        
        c_d1, c_d2 = st.columns(2)
        with c_d1:
            st.write(f"👉 **外資 CSV 原始文字檔前 500 字** ({test_tpex_csv})")
            f_txt = download_tpex_csv_text(test_tpex_csv, "qfiiStat")
            if f_txt:
                st.text(f_txt[:500])
            else:
                st.error("無資料或連線失敗")
        with c_d2:
            st.write(f"👉 **融資券 JSON 原始回傳** ({test_roc})")
            url_m = f"https://www.tpex.org.tw/web/stock/margin_trading/margin_balance/margin_bal_result.php?l=zh-tw&o=json&d={test_roc}"
            try:
                res_raw = requests.get(url_m, timeout=5, verify=False).text
                st.text(res_raw[:500])
            except Exception as e:
                st.error(f"解析失敗: {e}")
        
        st.info("💡 檢查完畢後，請取消勾選 Debug 模式，即可恢復正常繪圖。")
        st.stop()
    
    with st.spinner("⏳ 正在即時向證交所/櫃買中心調閱籌碼數據，請稍候..."):
        df_foreign_export, df_trust_export, df_margin_export, fig_f, fig_t = step6_extract_institutional_data(raw_ticker, valid_hist, is_otc)
        
    col_f, col_t = st.columns(2)
    with col_f: st.plotly_chart(fig_f, use_container_width=True)
    with col_t: st.plotly_chart(fig_t, use_container_width=True)
    
    if not df_margin_export.empty:
        st.markdown("#### 📊 近 10 日信用交易明細 (張)")
        col_m, col_s = st.columns(2)
        
        df_margin_reversed = df_margin_export.iloc[::-1].set_index('日期')
        
        with col_m:
            st.write("**💰 融資狀況 (散戶做多指標)**")
            st.dataframe(df_margin_reversed[['融資變動(張)', '融資餘額(張)']], use_container_width=True)
        with col_s:
            st.write("**📉 融券狀況 (散戶做空指標)**")
            st.dataframe(df_margin_reversed[['融券變動(張)', '融券餘額(張)']], use_container_width=True)

    st.divider()
    st.subheader("💾 匯出完整 Excel 報表")
    try:
        output = io.BytesIO()
        df_sr_excel = pd.DataFrame([{'項次': i+1, '價格級距區間 (TWD)': item['disp_label'], '累積成交量 (張)': int(item['vol'])} for i, item in enumerate(all_intervals_disp)])
        df_top5_excel = pd.DataFrame([{'位置': '⬆️ 向上壓力區', '價格級距區間': b['disp_label'], '累積成交量 (張)': int(b['vol'])} for b in top_5_above] + [{'位置': '🎯 最新股價', '價格級距區間': f"{current_price_round:.2f}", '累積成交量 (張)': 0}] + [{'位置': '⬇️ 向下支撐區', '價格級距區間': b['disp_label'], '累積成交量 (張)': int(b['vol'])} for b in top_5_below])
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_sr_excel.to_excel(writer, sheet_name='區間分價量總表', index=False)
            df_top5_excel.to_excel(writer, sheet_name='關鍵支撐壓力', index=False)
            
            if not df_foreign_export.empty: 
                df_foreign_export.to_excel(writer, sheet_name='外資買賣超(20日)', index=False)
            if not df_trust_export.empty:
                df_trust_export.to_excel(writer, sheet_name='投信買賣超(20日)', index=False)
            if not df_margin_export.empty:
                df_margin_export.to_excel(writer, sheet_name='融資券狀況(10日)', index=False)
            
            workbook = writer.book
            for ws in workbook.worksheets:
                for col in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 25.5
            
            sheet1 = workbook['區間分價量總表']
            chart = BarChart()
            chart.type, chart.style = "bar", 10
            chart.title = f"{target_name} {lookback}日分價量分佈圖"
            chart.x_axis.title, chart.y_axis.title = "價格區間", "成交量"
            chart.height, chart.width = max(10, len(df_sr_excel) * 0.5) * 1.5, 24
            chart.add_data(Reference(sheet1, min_col=3, min_row=1, max_row=len(df_sr_excel) + 1), titles_from_data=True)
            chart.set_categories(Reference(sheet1, min_col=2, min_row=2, max_row=len(df_sr_excel) + 1))
            sheet1.add_chart(chart, "E2")
            
        st.download_button("📥 點我下載 Excel 分析報表", data=output.getvalue(), file_name=f"{re.sub(r'[\\/*?:\"<>|]', '_', target_name)}_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
    except Exception as e:
        st.error(f"❌ Excel 產生錯誤：{e}")
